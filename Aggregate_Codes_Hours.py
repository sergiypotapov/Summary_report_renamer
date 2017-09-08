import openpyxl
import Get_hours_for_code

def Run_Excel_changes():
    wb = openpyxl.load_workbook("Luxoft_Timesheet_27_2017_Denys_LebedevV70.xlsx", data_only=True)
    ws = wb.active

    read_code_letter = 'C'
    read_code_digit = 7
    read_project_code = read_code_letter + str(read_code_digit)
    project_code_value = ws[read_project_code].value

    dict_code_hours = {}

#read Project Codes

    while project_code_value !=0:

        project_code_value = ws[read_project_code].value

        if project_code_value !=0:
            # Get hours for code
            total_hours_per_code = Get_hours_for_code.get_hours_for_code(read_code_letter, read_code_digit, ws)

            #put hours for
            read_code_digit = read_code_digit + 1
            read_project_code = read_code_letter + str (read_code_digit)

            if project_code_value in dict_code_hours:
                total_hours_per_code = dict_code_hours[project_code_value] + total_hours_per_code
                dict_code_hours[project_code_value] = total_hours_per_code
            else:
                dict_code_hours[project_code_value] = total_hours_per_code

            print(dict_code_hours)



Run_Excel_changes()