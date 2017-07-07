import openpyxl
import os

#next_month_name = "December"
#next_year_name = 2017


def Run_Excel_changes(full_file_path, list_of_first_end_day_in_week, next_year, next_month_name, next_month_name_short, next_year_short, next_file_path):
    wb = openpyxl.load_workbook(full_file_path, data_only=True)

    wb2 = openpyxl.load_workbook(full_file_path, data_only=False)

    sheet_list = wb.sheetnames
    print(sheet_list)
    iterator = 0
    print("Lists",len(sheet_list), len(list_of_first_end_day_in_week))
    for sheet in sheet_list:
        print('Sheet name',sheet)
        print ('iterator name',iterator)
        week = list_of_first_end_day_in_week[iterator]
        print("week",week)
        ws = wb2.get_sheet_by_name(sheet)
        ws.sheet_properties.tabColor = '40E0D0'
        to_sheet = str(week[0]) + ' - ' + str(week[1])
        #ws.title = to_sheet + ' ' + str(next_month_name_short) + ' ' +str(next_year_short)
        ws['A48'] = to_sheet + ' ' + str(next_month_name_short) + ' ' +str(next_year_short)
        iterator +=1

        ws['A2'].value = to_sheet + ' ' + next_month_name_short + ' ' +str(next_year)



        if iterator == len(list_of_first_end_day_in_week):
            print("CARAMBA")
            break
    print('RANAMING LAST SHEET')
    wb2.save('Pre_' + next_file_path)
    pre_file_name = 'Pre_' + next_file_path
    #last_sheet_ranamer(pre_file_name, next_month_name, next_year)

#def last_sheet_ranamer(pre_file_name, next_month_name, next_year):
    #wb = openpyxl.load_workbook(pre_file_name, data_only=True)
    #sheet_list = wb.sheetnames
    last_sheet_name = sheet_list[-1:]

    last_sheet_name = last_sheet_name[0]

    print('LAST SHEET NAME',last_sheet_name)

    ws = wb2.get_sheet_by_name(last_sheet_name)
    #ws.title = next_month_name + '_' + str(next_year)
    ws['A48'] =  next_month_name + '_' + str(next_year)
    ws.sheet_properties.tabColor = '6B8E23'


    ws['A2'].value = next_month_name + ' ' + str(next_year) + ' ' + '(total)'
    a = 'AUTO_' + pre_file_name
    wb2.save(a)
    print('SAVED TO....', 'AUTO_' + pre_file_name)

    print_new_sheet_list(a)

def print_new_sheet_list(a):
    wb = openpyxl.load_workbook(a, data_only=True)
    new_sheet_list = wb.sheetnames
    print(a, ' Contains next sheets: ', new_sheet_list)

def next_file_name(current_file_path, current_month_name, next_month_name, current_year, next_year):
        next_file_path = current_file_path.replace(current_month_name, next_month_name)
        next_year = str(next_year) ##2017
        current_year = str(current_year) ##2016
        next_file_path = next_file_path.replace(current_year, next_year)

        print(next_file_path)

        return next_file_path



