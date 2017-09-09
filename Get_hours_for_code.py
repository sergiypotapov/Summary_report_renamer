from getindex_fromstring_uppercase import get_index_from_letter
from getindex_fromstring_uppercase import get_letter_from_index
import string
import openpyxl

def get_hours_for_code( read_code_digit, ws, report_month):

    total_hours_per_code = 0
    ifint = True

    start_coordinate_letter_index, stop_coordinate_letter_index = compere_reported_month_toE1_and_H1(report_month, ws)
    stop_coordinate = stop_coordinate_letter_index + str(read_code_digit)
    start_coordinate = start_coordinate_letter_index + str(read_code_digit)

    while start_coordinate != stop_coordinate:

        start_coordinate = start_coordinate_letter_index + str(read_code_digit)


        read_code_value = ws[start_coordinate].value




        if ifint == isinstance(read_code_value, int):

            total_hours_per_code = total_hours_per_code + read_code_value

            index = get_index_from_letter(start_coordinate_letter_index) + 1

            read_code_letter = get_letter_from_index(index)
            start_coordinate_letter_index = read_code_letter



        else:

            index = get_index_from_letter(start_coordinate_letter_index) + 1

            read_code_letter = get_letter_from_index(index)
            start_coordinate_letter_index = read_code_letter



    return(total_hours_per_code)


def compere_reported_month_toE1_and_H1(report_month, ws):

    E1coordinate = str(ws['E1'].value)
    E1_month = E1coordinate[5:7]
    H1coordinate = str(ws['H1'].value)
    H1_month = H1coordinate[5:7]




    if report_month == E1_month and report_month == H1_month:
        print('Both')
        start_coordinate_letter_index = 'E'
        stop_coordinate_letter_index = 'K'


    elif report_month == E1_month:

        start_coordinate_letter_index = 'E'
        ore = 'left'
        stop_coordinate_letter_index = find_01_coordinte_letter(ws, ore)

    elif report_month == H1_month:
        
        ore = 'right'
        start_coordinate_letter_index = find_01_coordinte_letter(ws, ore)

        stop_coordinate_letter_index = 'K'

    else:
        print('The month is out of range', '\n Month under report: ', report_month, ' \n Monthes in file are: %s and %s' % (E1_month, H1_month))
        exit()
    return (start_coordinate_letter_index, stop_coordinate_letter_index)

def find_01_coordinte_letter(ws, ore):


    alph = 'EFGHIJKL'
    digit= '5'
    for letter in alph:

        cell = letter+digit

        value =  str(ws[cell].value)
        value = value[8:10]

        if value == '01':

            if ore == 'left':



                i = alph.index(letter) - 1
                letter = alph[i]

                break
            else:

                i = alph.index(letter)
                letter = alph[i]
                alph.index(letter)

                break

    return (letter)




